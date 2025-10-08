from collections.abc import Sequence
import pathlib
import re

from absl import app
from absl import flags
import immutabledict
import openpyxl
import yaml

from utils import logger_manager


_RESULT_INDEX = 0
_DETAILS_INDEX = 13
_SPLIT_PATTERN = "---\n"

# The sponge properties in the sheet.
# The key is the column name in the sponge properties.
# The value is the column number in the sheet.
_SPONGE_PROPERTIES_IN_SHEET = immutabledict.immutabledict({
    "SASS trigger switch": 1,
    "ACL connection initiated": 2,
    "profile connecting": 3,
    "ACL connection successful": 4,
    "profile connected": 5,
    "Active switch trigger": 6,
    "Receive active switch": 7,
    "HFP audio streaming": 8,
    "A2DP start streaming": 9,
    "State change latency": 10,
    "Audio change latency": 11,
    "debug_msg": 13,
})

_ISSUE_PRIORITY_REGEX = re.compile(r"\[(?P<priority>P\d) issue\]")

_STATE_TEST_CASE = (
    "test_1_01_most_recently_inuse_state",
    "test_1_02_no_streaming_state",
    "test_1_03_game_state",
    "test_1_04_media_state",
    "test_1_05_phonecallout_state",
    "test_1_06_voip_state",
)

_KEY_STATE_TEST_CASE = (
    "test_2_1_dut_a_inuse",
    "test_2_1_dut_a_mostrecentlyinuse",
    "test_2_2_dut_c_inuse",
    "test_2_3_dut_e_inuse",
    "test_2_4_from_dut_a_to_dut_n",
    "test_2_5",
    "test_2_6",
)

_KEY_STATE_MULTIPOINT_TEST_CASE = ("test_2_7_keystate", "test_2_8_keystate")

_E2E_TEST_CASE = (
    "test_3_01_phonecallout_not_switch_phonecallout",
    "test_3_02_phonecallout_not_switch_voip",
    "test_3_03_phonecallout_not_switch_media",
    "test_3_04_phonecallout_not_switch_game",
    "test_3_05_media_switch_phonecallout",
    "test_3_06_media_switch_voip",
    "test_3_07_media_switch_media",
    "test_3_08_media_not_switch_game",
    "test_3_09_game_switch_phonecallout",
    "test_3_10_game_switch_voip",
    "test_3_11_game_switch_media",
    "test_3_12_game_switch_game",
    "test_3_13_no_streaming_switch_phonecallout",
    "test_3_14_no_streaming_switch_voip",
    "test_3_15_no_streaming_switch_media",
    "test_3_16_no_streaming_switch_game",
    "test_3_17_revert_button",
    "test_3_18_revert_after_auto_connect",
    "test_3_19_auto_connect_via_phonecallout",
    "test_3_20_auto_connect_via_media",
    "test_3_21_non_sass_no_streaming_not_switch_media",
    "test_3_22_non_sass_no_streaming_not_switch_phonecallout",
    "test_3_23_auto_connect_via_phonecallout_if_lastest_is_non_sass",
    "test_3_24_auto_connect_via_media_if_lastest_is_non_sass",
    "test_3_25_media_not_switch_media_different_account",
    "test_3_26_no_streaming_not_switch_phonecallout_different_account",
    "test_3_27_auto_connect_via_media_different_accounts",
    "test_3_28_auto_connect_via_phonecallout_different_accounts",
    "test_3_29_revert_resume_playing",
)

_E2E_MULTIPOINT_TEST_CASE = (
    "test_3_30_phonecallout_not_switch_phonecallout",
    "test_3_31_phonecallout_not_switch_voip",
    "test_3_32_phonecallout_not_switch_media",
    "test_3_33_phonecallout_not_switch_game",
    "test_3_34_media_switch_phonecallout",
    "test_3_35_media_switch_voip",
    "test_3_36_media_switch_media",
    "test_3_37_media_not_switch_game",
    "test_3_38_game_switch_phonecallout",
    "test_3_39_game_switch_voip",
    "test_3_40_game_switch_media",
    "test_3_41_game_switch_game",
    "test_3_42_no_streaming_switch_phonecallout",
    "test_3_43_no_streaming_switch_voip",
    "test_3_44_no_streaming_switch_media",
    "test_3_45_no_streaming_switch_game",
    "test_3_46_revert_resume_playing",
    "test_3_47_media_switch_phonecallout_non_sass",
    "test_3_48_media_not_switch_media_non_sass",
    "test_3_49_no_streaming_switch_media_non_sass",
    "test_3_50_media_non_sass_switch_phonecallout",
    "test_3_51_media_non_sass_not_switch_media",
    "test_3_52_phonecallout_non_sass_not_switch_phonecallout_third",
    "test_3_53_phonecallout_non_sass_not_switch_media_third",
    "test_3_54",
    "test_3_55_mixed_devices",
    "test_3_56",
    "test_3_57_avail_flag",
)


def generate_row_dic() -> dict[str, int]:
  """Generate dictionary that records the row index of each test case.

  Returns:
    The dictionary that records the row index of each test case.
  """
  row_dict = {}

  # _STATE_TEST_CASE's first row index is 9 in external guideline template
  # - [Template] SASS Certification Automation Test Report:
  # https://docs.google.com/spreadsheets/d/1lgYFqUbRBmy_dxOqKpfyM37pAGCr6XA53Er-ahPmmh0/edit
  state_test_case_begin_row = 9
  for index, test_case in enumerate(_STATE_TEST_CASE):
    row_dict[test_case] = state_test_case_begin_row + index * 3

  key_state_test_case_begin_row = 37
  for index, test_case in enumerate(_KEY_STATE_TEST_CASE):
    row_dict[test_case] = key_state_test_case_begin_row + index * 3

  key_state_multipoint_test_case_begin_row = 59
  for index, test_case in enumerate(_KEY_STATE_MULTIPOINT_TEST_CASE):
    row_dict[test_case] = key_state_multipoint_test_case_begin_row + index * 3

  e2e_test_case_begin_row = 74
  for index, test_case in enumerate(_E2E_TEST_CASE):
    row_dict[test_case] = e2e_test_case_begin_row + index * 5

  e2e_multipoint_test_case_begin_row = 233
  for index, test_case in enumerate(_E2E_MULTIPOINT_TEST_CASE):
    row_dict[test_case] = e2e_multipoint_test_case_begin_row + index * 5

  return row_dict


def alph_to_num(alphabet: str) -> int:
  """Convert sheet column letter to corresponding numerical index.

  Args:
    alphabet: The chars from --column.

  Returns:
    The column number of the alphabet.
  """
  res = 0
  for ch in alphabet:
    res = res * 26
    res = res + ord(ch.upper()) - ord("A") + 1
  return res


def fill_cell(
    sht: openpyxl.worksheet.worksheet.Worksheet,
    row: int,
    column: int,
    value: str | float | None,
) -> None:
  """Fill the cell with the value.

  Args:
    sht: The sheet object.
    row: The row number.
    column: The column number.
    value: The value to fill.
  """
  if value in [None, 100, -100]:
    return
  logger_manager.logger.info(
      f"Fill cell: {row}, {column}, {value} ({type(value)})"
  )
  if sht.cell(row=row, column=column).value != "N/A":
    if isinstance(value, str) or value >= 0.0:
      sht.cell(row=row, column=column, value=value)


def get_cell_value(
    sht: openpyxl.worksheet.worksheet.Worksheet,
    row: int,
    column: int,
) -> str:
  """Get the cell value.

  Args:
    sht: The sheet object.
    row: The row number.
    column: The column number.

  Returns:
    The value of the cell.
  """
  return sht.cell(row=row, column=column).value


def get_priority(msg: str) -> str:
  """Get the priority from the error message.

  Args:
    msg: The message of the error log.

  Returns:
    The priority of the error. Like P0, P1, P2, P3, P4.
  """
  re_search = _ISSUE_PRIORITY_REGEX.search(msg)
  return re_search.group("priority") if re_search else "P0"


def get_test_name(test_name: str) -> tuple[str, int]:
  """Get the test name, test main and num from the test name.

  Args:
    test_name: The test name.

  Returns:
    The test name, test main and num.
  """

  sub_test_name = test_name
  num = 0

  # Get the test name and the number of sub test.
  # ex: test_1_01_most_recently_inuse_state_1
  # test_head = test_
  # test name: 1_01_most_recently_inuse_state
  # num: 1
  if re_find := re.fullmatch(
      r"(?P<test_head>test_)(?P<test_main>\d_\d{1,3}_\w+)(?P<num>_\d?)",
      test_name,
  ):
    test_head = re_find.group("test_head")
    test_main = re_find.group("test_main")
    num = re_find.group("num")
    sub_test_name = f"{test_head}{test_main}"
    num_s = num.replace("_", "").strip()
    num = int(num_s)

  return sub_test_name, num


def split_paragraphs(split_pattern: str, paragraphs: list[str]) -> list[str]:
  """Split the paragraph to small paragraphs.

  Args:
    split_pattern: The pattern to split.
    paragraphs: The paragraph to split.

  Returns:
    The list of small paragraphs.
  """
  if not paragraphs:
    return []

  if split_pattern not in paragraphs:
    return paragraphs

  # Split the string by split_pattern to get each test case.
  part_of_paragraphs = ""
  result = []
  # Make split rule to strictly match start with the pattern.
  for paragraph in paragraphs:
    # If the paragraph is the same as the split pattern, add the part of
    # paragraphs to the result.
    if re.fullmatch(split_pattern, paragraph):
      result.append(part_of_paragraphs)
      part_of_paragraphs = ""
      continue
    part_of_paragraphs += paragraph

  result.append(part_of_paragraphs)
  return [s for s in result if s]


def main(_: Sequence[str]):

  report_file_name = flag_xlsx_file_path.value
  # --sheet_name is a string flag of the sheet name.
  sheet_name = flag_xlsx_sheet_name.value
  # --column is a string flag of the column to start.
  # Default is "d".
  column_alphabet = flag_column.value
  # --yaml_file_path is a string flag of the yaml file path.
  yaml_file_name = flag_yaml_file_path.value
  # --debug_mode is a boolean flag of the debug mode.
  debug_mode = flag_debug_mode.value

  # This would be replace the original `logger_manager.py` instance.
  current_folder = pathlib.Path.cwd()
  output_folder = current_folder.joinpath("fill_report_logs")
  logger_manager.logger = (
      logger_manager.LoggerManager(name=__name__, debug_mode=debug_mode)
      .add_stream_handler()
      .add_file_handler(output_folder=output_folder)
      .get_logger()
  )

  report_file_path = pathlib.Path(report_file_name)
  yaml_file_path = pathlib.Path(yaml_file_name)

  logger_manager.logger.info(f"Report file path: {report_file_path}")

  # Convert the column alphabet to column number.
  # ex: A -> 1, B -> 2, C -> 3, AH -> 27, ...
  column_num = alph_to_num(column_alphabet)
  logger_manager.logger.info(f"Start column number: {column_num}")

  # Open the template excel file.
  wb: openpyxl.workbook.Workbook = openpyxl.load_workbook(report_file_path)
  # Get the sheet by name.
  sht = wb.get_sheet_by_name(name=sheet_name)

  logger_manager.logger.info(f"Open yaml file: {yaml_file_path}")

  row_dict = generate_row_dic()

  with open(yaml_file_name, "r") as yaml_file:
    lines = yaml_file.readlines()
    split_result = split_paragraphs(_SPLIT_PATTERN, lines)
    logger_manager.logger.info(f"Show split result: {split_result}")
    for paragraph in split_result:
      row = 0
      logger_manager.logger.info(
          "=======================paragraph=========================="
      )
      try:
        logger_manager.logger.info(f"Show paragraph: {paragraph}")
        parsed_dict = yaml.safe_load(paragraph)
        logger_manager.logger.info(
            "=========================parse dict==========================="
        )
        if isinstance(parsed_dict, dict):
          # test name is the key of the dict.
          # ex: test_1_01_most_recently_inuse_state
          test_name = parsed_dict.get("Test Name")

          # test_sp is the sponge properties.
          # ex. sponge_properties:
          #       A2DP start streaming: 100
          #       ACL connection initiated: 100
          #       ACL connection successful: 100
          #       Active switch trigger: 100
          #       Audio change latency: 3.028
          #       HFP audio streaming: 100
          #       Receive active switch: 0.777
          #       SASS trigger switch: 100
          #       State change latency: 0.849
          #       profile connected: 100
          #       profile connecting: 100
          test_sp = parsed_dict.get("sponge_properties")

          # test result is the test result.
          # ex: Result: PASS or FAIL or ERROR
          test_result = parsed_dict.get("Result")

          logger_manager.logger.info(f"Get test name: {test_name}")
          logger_manager.logger.info(f"Get test sp: {test_sp}")
          logger_manager.logger.info(f"Get test result: {test_result}")

          if test_name:
            logger_manager.logger.info(f"Show test name: `{test_name}`")
            sub_test_name, num = get_test_name(test_name)
            logger_manager.logger.info(
                f"Show sub test name: `{sub_test_name}`, num: `{num}`"
            )

            if sub_test_name not in row_dict.keys():
              # Skip the test case that no need to record.
              logger_manager.logger.info("Skip by sub test name.")
              continue

            row = row_dict[sub_test_name] + num
            logger_manager.logger.info(f"Get the test row: {row}")
            if test_result:
              # Clear the previous value if any.
              fill_cell(
                  sht, row=row, column=column_num + _RESULT_INDEX, value=""
              )
              fill_cell(
                  sht, row=row, column=column_num + _DETAILS_INDEX, value=""
              )

              if test_result == "PASS":
                # PASS = Y
                fill_cell(
                    sht, row=row, column=column_num + _RESULT_INDEX, value="Y"
                )
              elif test_result == "Skip":
                # Skip = N
                if (details := parsed_dict.get("Details")) is not None:
                  fill_cell(
                      sht,
                      row=row,
                      column=column_num + _DETAILS_INDEX,
                      value=details,
                  )
              else:
                # Get the issues priority!
                priority = "P2"

                # Handle error messages
                error_msg_set = set()

                if (details := parsed_dict.get("Details")) is not None:
                  error_msg_set.add(details)
                  priority = min(priority, get_priority(details))

                if (extras := parsed_dict.get("Extras")) is not None:
                  error_msg_set.add(extras)
                  priority = min(priority, get_priority(extras))

                extra_errors = parsed_dict.get("Extra Errors")
                for sub_dict in extra_errors.values():
                  if details := sub_dict.get("Details"):
                    error_msg_set.add(details)
                    priority = min(priority, get_priority(details))

                if error_msg_set:
                  debug_msg = "\n".join(list(error_msg_set))
                  fill_cell(
                      sht,
                      row=row,
                      column=column_num + _DETAILS_INDEX,
                      value=debug_msg,
                  )

                fill_cell(
                    sht,
                    row=row,
                    column=column_num + _RESULT_INDEX,
                    value=priority,
                )

            if test_sp:
              for key, value in _SPONGE_PROPERTIES_IN_SHEET.items():
                # Clear the previous value first.
                fill_cell(sht, row=row, column=column_num + value, value="")
                # Fill the value.
                # num 1 to 11 and 13 is for the test_sp.
                # num 12 is need to follow the tepmlate rule to use.
                fill_cell(
                    sht,
                    row=row,
                    column=column_num + value,
                    value=test_sp.get(key),
                )

      except yaml.YAMLError as exc:
        fill_cell(
            sht, row=row, column=column_num + _RESULT_INDEX, value="ERROR"
        )
        logger_manager.logger.error(f"Got Error: {exc}")
  wb.save(report_file_path)


if __name__ == "__main__":

  flag_xlsx_file_path = flags.DEFINE_string(
      name="xlsx_file",
      default=None,
      help="The xlsx file path.",
      required=True,
  )
  flag_xlsx_sheet_name = flags.DEFINE_string(
      name="sheet_name",
      default=None,
      help="The sheet name.",
      required=True,
  )
  flag_yaml_file_path = flags.DEFINE_string(
      name="yaml_file",
      default=None,
      help="The yaml file path.",
      required=True,
  )
  flag_column = flags.DEFINE_string(
      name="column",
      default=None,
      help=(
          "The column to start. In the form of alphabet, ex: [A, B, C, S, P, AH"
          " ...]"
      ),
      required=True,
  )
  flag_debug_mode = flags.DEFINE_bool(
      name="debug_mode",
      default=False,
      help="The debug mode of the fill report.",
      required=False,
  )

  # Make a flag required.
  flags.mark_flag_as_required("xlsx_file")
  flags.mark_flag_as_required("sheet_name")
  flags.mark_flag_as_required("yaml_file")
  flags.mark_flag_as_required("column")

  app.run(main)
